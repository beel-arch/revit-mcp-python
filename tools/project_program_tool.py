"""MCP tools: project program (programma van eisen) — opslag, raadpleging en
model-vergelijking.

programs/<project>.json is de uniforme vorm van elk programma, ongeacht de bron
(PvE-Excel, lokaalfiches, Room-by-Room matrix). Elke save schrijft ook een
leesbare .md-samenvatting ernaast. Zie programs/README.md voor het schema.
"""

import json
import re
from datetime import date
from pathlib import Path

from mcp.server.fastmcp import Context

from .project_profile_tool import load_current_profile, finest_grouping

PROGRAMS_DIR = Path(__file__).resolve().parent.parent / "programs"
EXPORTS_DIR = Path(__file__).resolve().parent.parent / "exports"


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _safe_key(key):
    key = re.sub(r'[<>:"/\\|?*\s]+', "-", (key or "").strip().lower()).strip("-")
    return key or "unnamed-project"


def program_path(project_key):
    return PROGRAMS_DIR / (_safe_key(project_key) + ".json")


def load_program(project_key):
    path = program_path(project_key)
    if not path.exists():
        return None
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return None


def _normalize_name(name):
    """Naam-join: lowercase, leestekens weg, whitespace samengevoegd."""
    n = re.sub(r"[()\[\]/\\.,;:'\"-]+", " ", (name or "").lower())
    return re.sub(r"\s+", " ", n).strip()


def _validate_program(program):
    errors = []
    if not isinstance(program, dict):
        return ["Programma moet een JSON-object zijn."]
    rooms = program.get("rooms")
    if not isinstance(rooms, list) or not rooms:
        errors.append("'rooms' moet een niet-lege lijst zijn.")
        return errors
    for i, r in enumerate(rooms):
        if not isinstance(r, dict):
            errors.append("rooms[{}] moet een object zijn.".format(i))
            continue
        if not r.get("name"):
            errors.append("rooms[{}] heeft een 'name' nodig.".format(i))
        for fld in ("count", "area_m2", "total_area_m2", "occupancy"):
            val = r.get(fld)
            if val is not None and not isinstance(val, (int, float)):
                errors.append("rooms[{}].{} moet numeriek zijn (of null).".format(i, fld))
    relations = program.get("relations")
    if relations is not None:
        if not isinstance(relations, list):
            errors.append("'relations' moet een lijst zijn.")
        else:
            for i, rel in enumerate(relations):
                if not (isinstance(rel, dict) and rel.get("a") and rel.get("b")):
                    errors.append("relations[{}] heeft 'a' en 'b' nodig.".format(i))
    return errors


def _program_md(program):
    """Leesbare .md-samenvatting van een programma."""
    rooms = program.get("rooms", [])
    lines = ["# Programma — {}".format(program.get("title") or program.get("project_name"))]
    lines.append("")
    lines.append("- Project key: `{}`".format(program.get("project_name")))
    if program.get("source_documents"):
        lines.append("- Bron: {}".format("; ".join(program["source_documents"])))
    if program.get("extracted"):
        lines.append("- Geëxtraheerd: {}".format(program["extracted"]))
    total = sum(r.get("total_area_m2") or 0 for r in rooms)
    n_instances = sum(r.get("count") or 0 for r in rooms)
    lines.append("- {} ruimtetypes, {} ruimte-instanties, {:.0f} m² totaal".format(
        len(rooms), int(n_instances), total))
    lines.append("")

    by_cluster = {}
    for r in rooms:
        by_cluster.setdefault(r.get("cluster") or "(geen cluster)", []).append(r)

    for cluster in sorted(by_cluster.keys()):
        lines.append("## {}".format(cluster))
        lines.append("")
        lines.append("| Code | Ruimte | Aantal | m²/ruimte | Totaal m² | Bezetting |")
        lines.append("|---|---|---|---|---|---|")
        for r in by_cluster[cluster]:
            lines.append("| {} | {} | {} | {} | {} | {} |".format(
                r.get("code") or "—", r.get("name") or "?",
                r.get("count") if r.get("count") is not None else "—",
                r.get("area_m2") if r.get("area_m2") is not None else "—",
                r.get("total_area_m2") if r.get("total_area_m2") is not None else "—",
                r.get("occupancy") if r.get("occupancy") is not None else "—"))
        lines.append("")

    relations = program.get("relations") or []
    if relations:
        lines.append("## Ruimtelijke relaties")
        lines.append("")
        for rel in relations:
            lines.append("- {} ↔ {}  ({}){}".format(
                rel.get("a"), rel.get("b"), rel.get("type") or "adjacent",
                "  — " + rel["note"] if rel.get("note") else ""))
        lines.append("")
    return "\n".join(lines)


def write_program(program, project_key):
    """Schrijf JSON + MD; geeft (json_path, md_path) terug. Ook gebruikt door scripts."""
    PROGRAMS_DIR.mkdir(parents=True, exist_ok=True)
    key = _safe_key(project_key)
    program["project_name"] = key
    json_path = PROGRAMS_DIR / (key + ".json")
    md_path = PROGRAMS_DIR / (key + ".md")
    json_path.write_text(
        json.dumps(program, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    md_path.write_text(_program_md(program) + "\n", encoding="utf-8")
    return json_path, md_path


# ---------------------------------------------------------------------------
# Vergelijking model ↔ programma
# ---------------------------------------------------------------------------

async def _resolve_program(revit_get, ctx, project_key=""):
    """(program, profile, link, foutmelding). Zonder key: via profiel van open model."""
    profile, model_name = await load_current_profile(revit_get, ctx)
    link = (profile or {}).get("program_link") or {}
    key = project_key.strip() if project_key else (link.get("program") or "")
    if link.get("mode") == "key_schedule" and not project_key:
        return None, profile, link, (
            "Dit project heeft program_link.mode = 'key_schedule': het programma zit "
            "ín het model. Gebruik get_rooms_overview met extra_parameters voor de "
            "eis- en verschilparameters (geen externe vergelijking nodig).")
    if not key:
        return None, profile, link, (
            "Geen programma gekoppeld aan model '{}'. Zet program_link in het profiel "
            "(save_project_profile) of geef project_key mee. Beschikbaar: {}".format(
                model_name or "?", ", ".join(
                    p.stem for p in PROGRAMS_DIR.glob("*.json")) or "geen"))
    program = load_program(key)
    if program is None:
        return None, profile, link, (
            "Programma '{}' niet gevonden in programs/. Run eerst de program-intake "
            "(zie skill) en save_project_program.".format(key))
    return program, profile, link, None


async def _run_comparison(revit_get, revit_post, ctx, project_key="", unit_filter="", level=""):
    """Structurele vergelijking. Geeft (rows, extras, meta) of (None, None, fout)."""
    program, profile, link, err = await _resolve_program(revit_get, ctx, project_key)
    if err:
        return None, None, err

    join_param = link.get("join_parameter") or ""
    group_param, _label = finest_grouping(profile)

    body = {
        "level": (level or "").strip(),
        "group_param": group_param,
        "group_prefix": (unit_filter or "").strip(),
        "params": [join_param] if join_param else [],
        "include_unplaced": False,
    }
    resp = await revit_post("/rooms/overview/", body, ctx)
    if isinstance(resp, str):
        return None, None, "Fout bij ophalen rooms: {}".format(resp)
    if isinstance(resp, dict) and "error" in resp:
        return None, None, "Fout bij ophalen rooms: {}".format(resp["error"])

    model_rooms = resp.get("rooms", [])

    # Index model rooms op join key
    index = {}
    for r in model_rooms:
        if join_param:
            key = ((r.get("params") or {}).get(join_param) or "").strip()
        else:
            key = _normalize_name(r.get("name"))
        if key:
            index.setdefault(key, []).append(r)

    rows = []
    matched_ids = set()
    for pr in program.get("rooms", []):
        if join_param:
            key = (pr.get("code") or "").strip()
        else:
            key = _normalize_name(pr.get("name"))
        matches = index.get(key, []) if key else []
        for m in matches:
            matched_ids.add(m.get("id"))

        expected = pr.get("count")
        min_area = pr.get("area_m2")
        actual = len(matches)
        areas = [m.get("area_m2") or 0 for m in matches]
        issues = []
        if expected is not None and actual != expected:
            issues.append("aantal: {} i.p.v. {}".format(actual, expected))
        if min_area is not None:
            too_small = [a for a in areas if a < min_area]
            if too_small:
                issues.append("{} ruimte(s) onder {} m² (kleinste: {} m²)".format(
                    len(too_small), min_area, min(too_small)))
        rows.append({
            "code": pr.get("code"),
            "name": pr.get("name"),
            "expected": expected,
            "actual": actual,
            "min_area": min_area,
            "areas": areas,
            "ok": not issues,
            "issues": issues,
        })

    extras = [r for r in model_rooms if r.get("id") not in matched_ids]
    meta = {
        "program": program,
        "join": join_param or "naam (genormaliseerd)",
        "unit_filter": (unit_filter or "").strip(),
        "level": (level or "").strip(),
        "model_room_count": len(model_rooms),
        "unplaced_skipped": resp.get("unplaced_skipped", 0),
    }
    return rows, extras, meta


def _comparison_report(rows, extras, meta):
    program = meta["program"]
    lines = ["# Programma-vergelijking — {}".format(
        program.get("title") or program.get("project_name"))]
    filt = []
    if meta["level"]:
        filt.append("level='{}'".format(meta["level"]))
    if meta["unit_filter"]:
        filt.append("unit='{}'".format(meta["unit_filter"]))
    lines.append("Join op: {}{}".format(meta["join"],
                 "  |  filter: " + ", ".join(filt) if filt else ""))
    if meta["unplaced_skipped"]:
        lines.append("⚠ {} unplaced rooms niet meegeteld".format(meta["unplaced_skipped"]))
    lines.append("")

    n_ok = sum(1 for r in rows if r["ok"])
    lines.append("## Resultaat: {}/{} programma-eisen OK".format(n_ok, len(rows)))
    lines.append("")
    lines.append("{:<8} {:<40} {:>7} {:>7} {:>9}  {}".format(
        "Code", "Ruimte", "Eis", "Model", "Min m²", "Status"))
    lines.append("-" * 90)
    for r in rows:
        status = "✓" if r["ok"] else "✗ " + "; ".join(r["issues"])
        lines.append("{:<8} {:<40} {:>7} {:>7} {:>9}  {}".format(
            (r["code"] or "—")[:8], (r["name"] or "?")[:40],
            r["expected"] if r["expected"] is not None else "—",
            r["actual"],
            r["min_area"] if r["min_area"] is not None else "—",
            status))
    lines.append("")

    if extras:
        lines.append("## Model-rooms zonder programma-match ({})".format(len(extras)))
        preview = {}
        for r in extras[:200]:
            n = r.get("name") or "?"
            preview[n] = preview.get(n, 0) + 1
        lines.append("  " + ", ".join(
            "{}×{}".format(c, n) if c > 1 else n
            for n, c in sorted(preview.items())))
        lines.append("  (circulatie/techniek hoort hier vaak terecht in te staan)")
    return "\n".join(lines)


# ---------------------------------------------------------------------------
# MCP tools
# ---------------------------------------------------------------------------

def register_project_program_tools(mcp, revit_get, revit_post):

    @mcp.tool()
    async def get_project_program(project_key: str = "", ctx: Context = None) -> str:
        """
        Toon het programma van eisen voor een project.

        Zonder project_key wordt het programma van het open model gebruikt
        (via program_link in het project profile). Het programma is ook los
        leesbaar: programs/<project>.md naast de JSON.

        Parameters:
          project_key: optioneel, bv. "snor-slod". Leeg = via profiel.
        """
        if project_key.strip():
            program = load_program(project_key)
            if program is None:
                available = ", ".join(p.stem for p in PROGRAMS_DIR.glob("*.json")) or "geen"
                return "Programma '{}' niet gevonden. Beschikbaar: {}".format(
                    project_key, available)
        else:
            program, _profile, _link, err = await _resolve_program(revit_get, ctx)
            if err:
                return err

        md = _program_md(program)
        path = program_path(program.get("project_name"))
        return "{}\n\n---\nBestand: {} (+ .md ernaast)".format(md, path)

    @mcp.tool()
    async def save_project_program(
        program_json: str,
        project_key: str,
        ctx: Context = None,
    ) -> str:
        """
        Valideer en bewaar een programma van eisen als programs/<project>.json
        + leesbare programs/<project>.md.

        Dit is de deliverable van de program-intake (zie de revit-query skill):
        de AI leest de brondocumenten (PvE-Excel, lokaalfiches, Room-by-Room
        matrix) EENMALIG, bouwt deze JSON, en slaat ze hier op. Daarna nooit
        meer de brondocumenten herlezen.

        Parameters:
          program_json: volledig programma als JSON-string. Schema (programs/README.md):
            {
              "title": "...",
              "source_documents": ["..."],
              "rooms": [
                {"code": "02.03", "name": "...", "cluster": "...",
                 "count": 2, "area_m2": 37.0, "total_area_m2": 74.0,
                 "occupancy": 12, "zone": "...", "use": "...",
                 "requirements": {}}
              ],
              "relations": [{"a": "02.03", "b": "13.01", "type": "adjacent"}]
            }
          project_key: korte projectsleutel, bv. "snor-slod" of "kuurne".
                       Koppel daarna in het model-profiel:
                       program_link = {"mode": "external_json", "program": "<key>"}.
        """
        try:
            program = json.loads(program_json)
        except Exception as e:
            return "Ongeldige JSON: {}".format(e)

        errors = _validate_program(program)
        if errors:
            return "Programma niet opgeslagen — validatiefouten:\n  - " + "\n  - ".join(errors)

        program.setdefault("extracted", date.today().isoformat())
        json_path, md_path = write_program(program, project_key)
        rooms = program.get("rooms", [])
        return (
            "Programma opgeslagen: {} ruimtetypes.\n"
            "  JSON: {}\n  Leesbaar: {}\n\n"
            "Koppel het aan een model via save_project_profile → "
            "program_link = {{\"mode\": \"external_json\", \"program\": \"{}\"}}"
        ).format(len(rooms), json_path, md_path, _safe_key(project_key))

    @mcp.tool()
    async def compare_model_with_program(
        ctx: Context,
        project_key: str = "",
        unit_filter: str = "",
        level: str = "",
    ) -> str:
        """
        Vergelijk het open Revit-model met het programma van eisen:
        aantallen en minimumoppervlaktes per ruimtetype (de "PvE-vergelijking").

        Join: via program_link.join_parameter uit het profiel (room-parameter met
        de programmacode), of anders op genormaliseerde roomnaam. Toont ook
        model-rooms zonder programma-match.

        Parameters:
          project_key: optioneel; leeg = programma via het profiel van het open model.
          unit_filter: optionele prefix op de grouping-waarde (lean filter in Revit).
          level:       optionele exacte levelnaam (lean filter in Revit).
        """
        rows, extras, meta = await _run_comparison(
            revit_get, revit_post, ctx, project_key, unit_filter, level)
        if rows is None:
            return meta
        return _comparison_report(rows, extras, meta)

    @mcp.tool()
    async def export_program_comparison(
        ctx: Context,
        project_key: str = "",
        unit_filter: str = "",
        level: str = "",
        file_name: str = "",
    ) -> str:
        """
        Schrijf de programma-vergelijking weg als Excel in verificatiematrix-stijl
        (kolommen: Code | Eistekst | Aantal eis | Aantal model | Min m² |
        Voldoet/Voldoet niet | Motivering). Bestand komt in exports/.

        Parameters: zoals compare_model_with_program, plus optionele file_name.
        """
        rows, _extras, meta = await _run_comparison(
            revit_get, revit_post, ctx, project_key, unit_filter, level)
        if rows is None:
            return meta

        import openpyxl
        from openpyxl.styles import Font

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Verificatie"
        headers = ["Code", "Eistekst", "Aantal eis", "Aantal model", "Min m²",
                   "Voldoet / Voldoet niet", "Motivering indien niet voldoet"]
        ws.append(headers)
        for c in ws[1]:
            c.font = Font(bold=True)
        for r in rows:
            eistekst = "{}× {} à min {} m²".format(
                r["expected"] if r["expected"] is not None else "?",
                r["name"],
                r["min_area"] if r["min_area"] is not None else "?")
            ws.append([
                r["code"] or "",
                eistekst,
                r["expected"],
                r["actual"],
                r["min_area"],
                "Voldoet" if r["ok"] else "Voldoet niet",
                "" if r["ok"] else "; ".join(r["issues"]),
            ])
        widths = [10, 55, 11, 13, 9, 20, 45]
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

        EXPORTS_DIR.mkdir(parents=True, exist_ok=True)
        program = meta["program"]
        name = file_name.strip() or "verificatie_{}_{}.xlsx".format(
            program.get("project_name"), date.today().isoformat())
        if not name.lower().endswith(".xlsx"):
            name += ".xlsx"
        path = EXPORTS_DIR / name
        wb.save(str(path))

        n_ok = sum(1 for r in rows if r["ok"])
        return "Verificatie-export geschreven: {}\n{}/{} eisen voldoen.".format(
            path, n_ok, len(rows))

    @mcp.tool()
    async def check_program_relations(
        ctx: Context,
        project_key: str = "",
        unit_filter: str = "",
    ) -> str:
        """
        Controleer de ruimtelijke relaties uit het programma (Schema Ruimtelijke
        Relaties) tegen de deur-adjacentiegraaf van het model: twee ruimtes zijn
        verbonden als ze een deur delen.

        Matching: programmacodes worden via het programma vertaald naar
        ruimtenamen en genormaliseerd vergeleken met de model-roomnamen.

        Parameters:
          project_key: optioneel; leeg = via profiel.
          unit_filter: optionele prefix op de grouping-waarde.
        """
        program, _profile, _link, err = await _resolve_program(revit_get, ctx, project_key)
        if err:
            return err
        relations = program.get("relations") or []
        if not relations:
            return ("Programma '{}' bevat geen ruimtelijke relaties. Voeg ze toe via "
                    "program-intake (relations: [{{\"a\": ..., \"b\": ...}}]).").format(
                        program.get("project_name"))

        # Rooms + deuren ophalen
        from urllib.parse import quote
        from .project_profile_tool import get_grouping, DEFAULT_GROUPING_PARAMETER
        group_param, _lbl, _prof = await get_grouping(revit_get, ctx)
        if unit_filter and group_param == DEFAULT_GROUPING_PARAMETER:
            endpoint = "/rooms/with_doors/apartment/{}".format(quote(unit_filter, safe=""))
        elif group_param != DEFAULT_GROUPING_PARAMETER:
            endpoint = "/rooms/with_doors/groupby/{}/{}".format(
                quote(group_param, safe=""),
                quote(unit_filter, safe="") if unit_filter else "all")
        else:
            endpoint = "/rooms/with_doors/"
        resp = await revit_get(endpoint, ctx)
        if isinstance(resp, str):
            return "Fout bij ophalen deuren: {}".format(resp)
        if isinstance(resp, dict) and "error" in resp:
            return "Fout bij ophalen deuren: {}".format(resp["error"])

        # Adjacentie: rooms die een door element_id delen
        door_rooms = {}
        name_by_norm = {}
        for room in resp.get("rooms", []):
            norm = _normalize_name(room.get("room_name"))
            name_by_norm.setdefault(norm, room.get("room_name"))
            for d in room.get("doors", []):
                door_rooms.setdefault(d.get("element_id"), set()).add(norm)
        adjacency = {}
        for members in door_rooms.values():
            for a in members:
                for b in members:
                    if a != b:
                        adjacency.setdefault(a, set()).add(b)

        # Code → programmanaam
        name_by_code = {}
        for pr in program.get("rooms", []):
            if pr.get("code"):
                name_by_code[pr["code"]] = pr.get("name")

        def resolve(ref):
            """Relatie-referentie (code of naam) → genormaliseerde naam."""
            return _normalize_name(name_by_code.get(ref, ref))

        lines = ["# Ruimtelijke relaties — {}".format(
            program.get("title") or program.get("project_name")), ""]
        n_ok = 0
        for rel in relations:
            a, b = resolve(rel.get("a")), resolve(rel.get("b"))
            a_found, b_found = a in adjacency or a in name_by_norm, b in adjacency or b in name_by_norm
            connected = b in adjacency.get(a, set())
            if connected:
                n_ok += 1
                status = "✓ verbonden (gedeelde deur)"
            elif not a_found or not b_found:
                missing = [rel.get("a") if not a_found else None,
                           rel.get("b") if not b_found else None]
                status = "? ruimte niet gevonden in model: {}".format(
                    ", ".join(m for m in missing if m))
            else:
                status = "✗ geen gedeelde deur gevonden"
            lines.append("  {} ↔ {}  — {}".format(rel.get("a"), rel.get("b"), status))
        lines.append("")
        lines.append("Resultaat: {}/{} relaties bevestigd.".format(n_ok, len(relations)))
        lines.append("Let op: relaties via open verbindingen (zonder deur) zijn niet "
                     "detecteerbaar via de deurgraaf.")
        return "\n".join(lines)
